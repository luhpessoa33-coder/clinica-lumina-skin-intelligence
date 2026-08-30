import csv
import json
from collections import defaultdict
from datetime import date
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1]
OUT = BASE / "docs" / "dados"
ART = BASE / "docs" / "graficos"
OUT.mkdir(parents=True, exist_ok=True)
ART.mkdir(parents=True, exist_ok=True)

items = []

def add(category, item, unit, qty, essential, recommended, premium, phase, note=""):
    items.append({
        "id": len(items) + 1, "categoria": category, "item": item, "unidade": unit,
        "quantidade": qty, "essencial": essential, "recomendado": recommended,
        "premium": premium, "fase": phase, "observacao": note
    })

# Projetos, licenças e implantação documental.
add("Projetos e licenças", "Levantamento topográfico/cadastral e sondagem simplificada", "vb", 1, 3500, 6000, 9000, "Antes da obra", "Atualizar após cidade e terreno.")
add("Projetos e licenças", "Projeto arquitetônico legal e executivo", "vb", 1, 12000, 22000, 35000, "Antes da obra", "Contratar arquiteto local com RRT.")
add("Projetos e licenças", "Projetos estrutural, elétrico, hidráulico, climatização e prevenção", "vb", 1, 8000, 15000, 25000, "Antes da obra", "Responsáveis com ART/RRT.")
add("Projetos e licenças", "Aprovação sanitária, prefeitura, bombeiros e taxas", "vb", 1, 3000, 7000, 12000, "Antes da obra", "Varia por município e estado.")
add("Projetos e licenças", "PGRSS, POPs, documentação técnica e implantação LGPD", "vb", 1, 2500, 5000, 8000, "Pré-operação", "Escopo final pelo RT.")
add("Projetos e licenças", "Abertura, contabilidade, contratos e seguros iniciais", "vb", 1, 3500, 7000, 12000, "Pré-operação", "Não inclui tributos mensais.")
add("Projetos e licenças", "Cursos complementares e treinamento de equipamentos", "vb", 1, 5000, 10000, 20000, "2026–2027", "Além da especialização já em curso.")

# Construção do módulo novo de 60 m².
add("Construção 60 m²", "Canteiro, proteção da casa e instalações provisórias", "vb", 1, 6000, 9000, 14000, "Obra", "Inclui isolamento e controle de poeira.")
add("Construção 60 m²", "Movimento de terra, fundações e impermeabilização de base", "vb", 1, 26000, 38000, 55000, "Obra", "Alto risco até sondagem e níveis reais.")
add("Construção 60 m²", "Estrutura, lajes/vergas e alvenarias", "vb", 1, 42000, 58000, 78000, "Obra", "Módulo 6 × 10 m.")
add("Construção 60 m²", "Cobertura, calhas, rufos e impermeabilização", "vb", 1, 18000, 26000, 38000, "Obra", "Destino pluvial depende do lote.")
add("Construção 60 m²", "Revestimentos de paredes e preparação de superfícies", "vb", 1, 12000, 18000, 26000, "Obra", "Superfícies lisas, íntegras e laváveis.")
add("Construção 60 m²", "Pisos, rodapés e revestimentos molhados", "m²", 60, 18000, 30000, 48000, "Obra", "Total instalado; inclui perdas e soleiras.")
add("Construção 60 m²", "Portas, janelas, ferragens e vidros", "vb", 1, 12000, 22000, 36000, "Obra", "Conferir esquadrias J01–J07 e P01–P07.")
add("Construção 60 m²", "Forros, pintura e acabamentos finais", "vb", 1, 14000, 22000, 34000, "Obra", "Tintas laváveis de baixo odor.")
add("Construção 60 m²", "Instalações hidráulicas, esgoto, louças e metais", "vb", 1, 14000, 22000, 35000, "Obra", "Inclui pontos clínicos e área técnica.")
add("Construção 60 m²", "Instalações elétricas, quadros, dados e aterramento", "vb", 1, 16000, 26000, 40000, "Obra", "Dimensionar cargas reais dos equipamentos.")
add("Construção 60 m²", "Climatização, renovação/exaustão e drenos", "vb", 1, 12000, 18000, 28000, "Obra", "Projeto específico e equipamentos eficientes.")
add("Construção 60 m²", "Sanitário acessível completo", "vb", 1, 7000, 12000, 20000, "Obra", "Inclui barras, reforços e acessórios.")
add("Construção 60 m²", "Segurança, emergência e sinalização técnica", "vb", 1, 2500, 5000, 9000, "Obra", "Validar com Bombeiros.")
add("Construção 60 m²", "Testes, comissionamento, as built e limpeza pós-obra", "vb", 1, 2500, 4500, 7000, "Entrega", "Condicionar pagamento final aos testes.")

# Área externa-base de 30 m².
add("Área externa 30 m²", "Escavação, regularização e retirada de material", "vb", 1, 3000, 5000, 8000, "Obra", "Ajustar aos níveis reais.")
add("Área externa 30 m²", "Base granular drenante e geotêxtil", "vb", 1, 3000, 5500, 8000, "Obra", "Conforme prancha D-01.")
add("Área externa 30 m²", "Dreno linear, tubos e caixas de inspeção", "vb", 1, 3500, 6500, 10000, "Obra", "Destino legal a confirmar.")
add("Área externa 30 m²", "Rota acessível em piso permeável antiderrapante", "m²", 12.96, 2500, 4500, 7000, "Obra", "Inclui 8% de perdas.")
add("Área externa 30 m²", "Grama sintética permeável 30–35 mm instalada", "m²", 19.44, 2000, 3500, 5500, "Obra", "Exigir vazão, UV e garantia.")
add("Área externa 30 m²", "Bordas, contenções e acabamentos", "vb", 1, 1500, 2500, 4000, "Obra", "Evitar degraus e aprisionamento de água.")
add("Área externa 30 m²", "Iluminação externa e paisagismo de baixa manutenção", "vb", 1, 1500, 3500, 7000, "Final", "Sem bloquear drenagem.")
add("Área externa 30 m²", "Mão de obra especializada e testes de drenagem", "vb", 1, 5000, 8000, 12000, "Obra", "Teste de chuva antes do aceite.")

# Mobiliário, conforto e tecnologia.
add("Mobiliário e interiores", "Marcenaria clínica e área técnica", "vb", 1, 12000, 22000, 35000, "Obra", "Armários fechados, bancada, cuba e estoque.")
add("Mobiliário e interiores", "Maca elétrica 2 motores", "un", 1, 3500, 5500, 7000, "Fase 1", "Capacidade e garantia por escrito.")
add("Mobiliário e interiores", "Mocho sela ergonômico", "un", 2, 800, 1200, 1700, "Fase 1", "Dois profissionais/apoios.")
add("Mobiliário e interiores", "Carrinhos auxiliares", "un", 2, 800, 1200, 1800, "Fase 1", "Separar material limpo.")
add("Mobiliário e interiores", "Lupa luminária", "un", 1, 350, 600, 800, "Fase 1", "Base estável e lavável.")
add("Mobiliário e interiores", "Balcão acessível e mobiliário administrativo", "vb", 1, 4000, 8000, 14000, "Obra", "Trecho acessível conforme A-03.")
add("Mobiliário e interiores", "Recepção completa para quatro pessoas", "vb", 1, 2300, 5000, 10000, "Fase 1", "Assentos, mesa, filtro, café, persiana, Wi-Fi e acessórios.")
add("Mobiliário e interiores", "Fotografia padronizada", "kit", 1, 1000, 2000, 3350, "Fase 1", "Tripé, luz difusa, fundo e escala.")
add("Mobiliário e interiores", "Comunicação visual, placas e decoração", "vb", 1, 1200, 3000, 6000, "Final", "Inclui placas de biossegurança.")
add("Mobiliário e interiores", "Apoio/copa e eletroportáteis", "vb", 1, 1000, 2500, 5000, "Fase 1", "Sem alimentos na área técnica.")

add("Tecnologia e segurança", "Notebook administrativo", "un", 1, 3000, 4500, 5500, "Fase 1", "Criptografia e bloqueio automático.")
add("Tecnologia e segurança", "Impressora multifuncional", "un", 1, 900, 1200, 1800, "Fase 1", "Posição protegida do público.")
add("Tecnologia e segurança", "Impressora de etiquetas", "un", 1, 0, 700, 1200, "Fase 1", "Rastreabilidade de estoque.")
add("Tecnologia e segurança", "Roteador, rede separada e nobreak", "kit", 1, 1100, 1700, 2300, "Fase 1", "Rede de visitantes separada.")
add("Tecnologia e segurança", "Câmeras, alarme e controle de acesso", "kit", 1, 1500, 3000, 5000, "Fase 1", "Nunca posicionar câmera na sala de procedimentos.")
add("Tecnologia e segurança", "Domínio, ferramentas e serviços digitais do primeiro ano", "vb", 1, 1000, 2500, 5000, "Operação", "Revisar assinaturas antes de renovar.")

# Equipamentos: cada linha escolhe uma alternativa por cenário.
add("Equipamentos", "HIFU selecionado: Sonofocus / Ultrafocus / Ultramed", "un", 1, 8091, 10990, 106656, "Fase 2", "Uma alternativa por cenário; não somar os três.")
add("Equipamentos", "Fotobiomodulação: Laserpulse / Fluence / Antares", "un", 1, 1980, 2390, 4474, "Fase 2", "Uma alternativa por cenário.")
add("Equipamentos", "Limpeza de pele: alta frequência / Sonopeel / plataforma", "un", 1, 500, 3300, 6490, "Fase 1/2", "Uma configuração por cenário.")
add("Equipamentos", "Analisador de pele: portátil / intermediário / avançado", "un", 1, 269, 350, 5000, "Fase 1/expansão", "Não substitui avaliação profissional.")
add("Equipamentos", "Radiofrequência HTM Effect", "un", 1, 0, 0, 8690, "Expansão", "Adiar no recomendado para preservar caixa.")
add("Equipamentos", "Autoclave e seladora", "kit", 1, 0, 4059, 8500, "Condicional", "Zero se processamento for terceirizado.")
add("Equipamentos", "Câmara de conservação e monitoramento", "kit", 1, 0, 8000, 9650, "Condicional", "Reservar, mas comprar somente após habilitação/licença.")
add("Equipamentos", "Sinais vitais e apoio clínico", "kit", 1, 500, 1200, 1800, "Fase 1", "Seleção e calibração pelo RT.")
add("Equipamentos", "Manutenção, calibração e acessórios iniciais", "vb", 1, 1500, 3500, 8000, "Fase 1/2", "Contrato e reserva técnica.")

# Estoques e consumíveis.
add("Estoque inicial", "Materiais gerais, EPIs e biossegurança", "vb", 1, 2500, 4500, 6500, "Fase 1", "Escala de 40 atendimentos/mês.")
add("Estoque inicial", "Linha profissional não invasiva", "vb", 1, 1000, 2500, 5000, "Fase 1", "Bioage/Tulípia/ADCOS conforme cadastro.")
add("Estoque inicial", "Consumíveis de aparelhos", "vb", 1, 500, 1500, 3000, "Fase 2", "Comprar próximo ao comissionamento.")
add("Estoque inicial", "Consumíveis de esterilização", "vb", 1, 0, 1000, 2500, "Condicional", "Zero se processamento terceirizado.")
add("Estoque inicial", "Reserva para injetáveis piloto", "vb", 1, 0, 8000, 20000, "Após licença", "Reservar caixa; não comprar antecipadamente.")
add("Estoque inicial", "Kit de emergência definido pelo RT", "vb", 1, 0, 2000, 4000, "Após licença", "Conteúdo somente após protocolo aprovado.")
add("Estoque inicial", "Materiais de limpeza e sanitário", "vb", 1, 1000, 1500, 2500, "Fase 1", "Inclui mops, panos e reposição.")
add("Estoque inicial", "Água, café e itens de recepção", "vb", 1, 500, 1000, 1500, "Fase 1", "Primeiro ciclo de abastecimento.")

# Pré-operação e caixa.
add("Lançamento e operação", "Identidade aplicada, inauguração e comunicação inicial", "vb", 1, 3000, 8000, 15000, "Pré-abertura", "Sem publicidade irregular ou promessa de resultado.")
add("Lançamento e operação", "Contrato/coleta de resíduos e implantação", "vb", 1, 1000, 2500, 4000, "Pré-abertura", "Fornecedor licenciado.")
add("Lançamento e operação", "Contratos iniciais de manutenção e suporte", "vb", 1, 1500, 3500, 8000, "Pré-abertura", "Equipamentos, climatização e segurança.")
add("Capital de giro", "Reserva operacional de três meses", "vb", 1, 30000, 60000, 90000, "Antes da abertura", "Pessoal, consumo, utilidades, taxas e marketing; recalcular com despesas reais.")

scenarios = ["essencial", "recomendado", "premium"]
rates = {
    "Projetos e licenças": {"essencial": 0.15, "recomendado": 0.20, "premium": 0.25},
    "Construção 60 m²": {"essencial": 0.20, "recomendado": 0.22, "premium": 0.25},
    "Área externa 30 m²": {"essencial": 0.20, "recomendado": 0.22, "premium": 0.25},
    "Mobiliário e interiores": {"essencial": 0.15, "recomendado": 0.18, "premium": 0.20},
    "Tecnologia e segurança": {"essencial": 0.12, "recomendado": 0.15, "premium": 0.18},
    "Equipamentos": {"essencial": 0.12, "recomendado": 0.15, "premium": 0.18},
    "Estoque inicial": {"essencial": 0.10, "recomendado": 0.12, "premium": 0.15},
    "Lançamento e operação": {"essencial": 0.10, "recomendado": 0.15, "premium": 0.20},
    "Capital de giro": {"essencial": 0.00, "recomendado": 0.00, "premium": 0.00},
}
escalation_rate = 0.05

category_totals = {s: defaultdict(float) for s in scenarios}
for row in items:
    for s in scenarios:
        category_totals[s][row["categoria"]] += float(row[s])

summary = {}
for s in scenarios:
    base_without_working = sum(v for c, v in category_totals[s].items() if c != "Capital de giro")
    working = category_totals[s]["Capital de giro"]
    escalation = base_without_working * escalation_rate
    contingency = sum(value * rates[cat][s] for cat, value in category_totals[s].items() if cat != "Capital de giro")
    total = base_without_working + escalation + contingency + working
    summary[s] = {
        "base_sem_capital_giro": round(base_without_working, 2),
        "reajuste_ate_maio_2027_5pct": round(escalation, 2),
        "contingencia": round(contingency, 2),
        "capital_giro": round(working, 2),
        "total_geral": round(total, 2),
        "reserva_mensal_8_meses": round(total / 8, 2),
    }

csv_path = OUT / "orcamento_detalhado.csv"
json_path = OUT / "orcamento_detalhado.json"
with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=list(items[0].keys()))
    w.writeheader(); w.writerows(items)
json_path.write_text(json.dumps({"itens": items, "resumo": summary, "taxas_contingencia": rates, "reajuste": escalation_rate}, ensure_ascii=False, indent=2), encoding="utf-8")

def brl(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

md = [
    "# Orçamento detalhado de implantação", "", "**Clínica Lumina Skin Intelligence**  ",
    "**Base:** módulo novo de 60 m² + área externa drenante de 30 m²  ",
    "**Data-base:** agosto de 2026 | **Início físico:** maio de 2027", "",
    "> Este orçamento é paramétrico porque cidade/UF, topografia, solo, conexão com a casa e cotações locais ainda não foram informados. A folga não é lucro do fornecedor: reúne reajuste até maio de 2027 e contingência técnica explícita. Antes de contratar, exigir no mínimo três propostas comparáveis e orçamento executivo do profissional local.", "",
    "## Total geral com folga", "",
    "| Cenário | Base sem capital de giro | Reajuste 5% | Contingência | Capital de giro | **Total geral** | Reserva mensal em 8 meses |",
    "|---|---:|---:|---:|---:|---:|---:|",
]
for s in scenarios:
    x = summary[s]
    md.append(f"| {s.title()} | {brl(x['base_sem_capital_giro'])} | {brl(x['reajuste_ate_maio_2027_5pct'])} | {brl(x['contingencia'])} | {brl(x['capital_giro'])} | **{brl(x['total_geral'])}** | {brl(x['reserva_mensal_8_meses'])} |")

md += ["", "## Subtotais por categoria", "", "| Categoria | Essencial | Recomendado | Premium |", "|---|---:|---:|---:|"]
for cat in rates:
    md.append(f"| {cat} | {brl(category_totals['essencial'][cat])} | {brl(category_totals['recomendado'][cat])} | {brl(category_totals['premium'][cat])} |")

md += ["", "## Orçamento linha a linha", "", "| Categoria | Item | Qtd. | Essencial | Recomendado | Premium | Fase |", "|---|---|---:|---:|---:|---:|---|"]
for r in items:
    md.append(f"| {r['categoria']} | {r['item']} | {r['quantidade']} {r['unidade']} | {brl(r['essencial'])} | {brl(r['recomendado'])} | {brl(r['premium'])} | {r['fase']} |")

md += [
    "", "## Interpretação dos cenários", "",
    "O cenário **Essencial** preserva estrutura e segurança, adia radiofrequência, cadeia fria e processamento interno e escolhe equipamentos de entrada. O cenário **Recomendado** é o orçamento-base com folga solicitado: inclui acabamento clínico consistente, Ultrafocus como referência HIFU, fotobiomodulação intermediária, reserva para cadeia fria e injetáveis sem antecipar a compra. O cenário **Premium** incorpora Ultramed HIFU, maior marcenaria, mais tecnologia e acabamentos superiores.", "",
    "A reserva mensal pressupõe oito aportes entre setembro de 2026 e abril de 2027. Se a clínica já possuir parte do capital, o painel recalculará a meta como `(total − saldo atual) ÷ meses restantes`. Compra de equipamentos, cosméticos e injetáveis deve ocorrer perto do comissionamento para preservar garantia e validade.", "",
    "## Fontes de calibração", "",
    "O SINAPI nacional de maio de 2026 registrou R$ 1.953,08/m², sendo R$ 1.104,59 de materiais e R$ 848,49 de mão de obra.[1] O CUB R8-N sem desoneração de São Paulo foi R$ 2.231,37/m² em julho de 2026.[2] Esses indicadores não representam sozinhos uma clínica pronta: o orçamento adiciona instalações, acessibilidade, acabamento sanitário, projetos, equipamentos, mobiliário, documentação, contingência e capital de giro.", "",
    "![Comparativo dos cenários](graficos/orcamento_cenarios.png)", "",
    "## Referências", "",
    "[1]: https://agenciadenoticias.ibge.gov.br/agencia-noticias/2012-agencia-de-noticias/noticias/47106-custos-da-construcao-variam-0-36-em-maio-com-destaque-para-regiao-sul \"IBGE — SINAPI maio de 2026\"", "[2]: https://sindusconsp.com.br/servicos/cub/ \"SindusCon-SP — CUB julho de 2026\"",
]
(BASE / "docs" / "orcamento_detalhado.md").write_text("\n".join(md) + "\n", encoding="utf-8")

# Planilha editável.
wb = Workbook()
ws = wb.active
ws.title = "Resumo"
headers = ["Cenário", "Base sem capital de giro", "Reajuste 5%", "Contingência", "Capital de giro", "Total geral", "Reserva mensal 8 meses"]
ws.append(headers)
for s in scenarios:
    x = summary[s]
    ws.append([s.title(), x["base_sem_capital_giro"], x["reajuste_ate_maio_2027_5pct"], x["contingencia"], x["capital_giro"], x["total_geral"], x["reserva_mensal_8_meses"]])

wd = wb.create_sheet("Detalhamento")
detail_headers = ["ID", "Categoria", "Item", "Unidade", "Quantidade", "Essencial", "Recomendado", "Premium", "Fase", "Observação"]
wd.append(detail_headers)
for r in items:
    wd.append([r["id"], r["categoria"], r["item"], r["unidade"], r["quantidade"], r["essencial"], r["recomendado"], r["premium"], r["fase"], r["observacao"]])

wc = wb.create_sheet("Contingências")
wc.append(["Categoria", "Essencial", "Recomendado", "Premium"])
for cat, vals in rates.items():
    wc.append([cat, vals["essencial"], vals["recomendado"], vals["premium"]])

wr = wb.create_sheet("Plano de reserva")
wr.append(["Parâmetro", "Valor"])
wr.append(["Meta recomendada com folga", summary["recomendado"]["total_geral"]])
wr.append(["Saldo já reservado — editar", 0])
wr.append(["Meses restantes — editar", 8])
wr.append(["Aporte mensal necessário", "=MAX(0,(B2-B3)/B4)"])
wr.append([])
wr.append(["Mês", "Aporte planejado", "Acumulado projetado"])
months = ["Set/2026", "Out/2026", "Nov/2026", "Dez/2026", "Jan/2027", "Fev/2027", "Mar/2027", "Abr/2027"]
for idx, month in enumerate(months, start=8):
    wr.cell(idx, 1, month)
    wr.cell(idx, 2, "=$B$5")
    wr.cell(idx, 3, f"=$B$3+SUM($B$8:B{idx})")
wr["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
wr["B4"].fill = PatternFill("solid", fgColor="FFF2CC")

wp = wb.create_sheet("Plano de compras")
wp.append(["Período", "Decisão/compra", "Fazer", "Evitar"])
purchase_plan = [
    ("Set–Nov/2026", "Levantamentos, anteprojeto e validação profissional", "Medir terreno, definir RT, consultar Vigilância e formar reserva", "Comprar aparelhos, cosméticos ou injetáveis"),
    ("Dez/2026–Fev/2027", "Projetos executivos e aprovações", "Contratar responsáveis, fechar projetos e três cotações por pacote", "Fechar obra sem escopo e cronograma"),
    ("Mar–Abr/2027", "Contratação da obra", "Assinar contrato por etapas, seguro e garantias", "Pagar grande adiantamento sem medição"),
    ("Mai–Ago/2027", "Obra e instalações", "Comprar materiais conforme cronograma e conferir entregas", "Levar equipamentos para ambiente com poeira"),
    ("Set–Out/2027", "Mobiliário, TI e equipamentos essenciais", "Instalar, testar, treinar e registrar garantias", "Comprar alternativas redundantes"),
    ("Nov–Dez/2027", "Estoque inicial e licenças finais", "Comprar consumíveis perto da abertura e validar lotes", "Antecipar injetáveis sem licença/agendas"),
    ("Pós-abertura", "Expansão", "Comprar RF, análise avançada e premium após demanda comprovada", "Imobilizar caixa por imagem/status"),
]
for row in purchase_plan:
    wp.append(row)

for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="254B45")
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for col in range(1, sheet.max_column + 1):
        max_len = min(60, max(len(str(sheet.cell(row=r, column=col).value or "")) for r in range(1, sheet.max_row + 1)) + 2)
        sheet.column_dimensions[get_column_letter(col)].width = max_len
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if sheet.title == "Resumo" and isinstance(cell.value, (int, float)) and cell.column >= 2:
                cell.number_format = 'R$ #,##0.00'
            elif sheet.title == "Detalhamento" and isinstance(cell.value, (int, float)) and cell.column in (6, 7, 8):
                cell.number_format = 'R$ #,##0.00'
            elif sheet.title == "Contingências" and isinstance(cell.value, (int, float)) and cell.column in (2, 3, 4):
                cell.number_format = '0%'
            elif sheet.title == "Plano de reserva" and isinstance(cell.value, (int, float)) and cell.column in (2, 3):
                cell.number_format = 'R$ #,##0.00'

xlsx_path = BASE / "docs" / "Orcamento_Clinica_Lumina.xlsx"
wb.save(xlsx_path)

# Gráfico por categoria, incluindo capital de giro.
cats = list(rates.keys())
x = range(len(cats))
width = 0.25
fig, ax = plt.subplots(figsize=(14, 7))
colors = ["#A5B8A6", "#41685F", "#C59B5A"]
for idx, s in enumerate(scenarios):
    vals = [category_totals[s][c] for c in cats]
    ax.bar([i + (idx - 1) * width for i in x], vals, width=width, label=s.title(), color=colors[idx])
ax.set_xticks(list(x))
ax.set_xticklabels(cats, rotation=28, ha="right")
ax.set_ylabel("Valor-base (R$)")
ax.set_title("Clínica Lumina Skin Intelligence — orçamento-base por categoria")
ax.legend(frameon=False)
ax.spines[["top", "right"]].set_visible(False)
fig.tight_layout()
fig.savefig(ART / "orcamento_cenarios.png", dpi=180, facecolor="white")
plt.close(fig)

print(json.dumps(summary, ensure_ascii=False, indent=2))
