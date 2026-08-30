from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

path = Path(__file__).resolve().parents[1] / "docs" / "Orcamento_Clinica_Lumina.xlsx"
wb = load_workbook(path)
if "Cronograma 18 meses" in wb.sheetnames:
    del wb["Cronograma 18 meses"]
ws = wb.create_sheet("Cronograma 18 meses")
headers = ["Mês", "Período", "Fase", "% do total", "60 m²", "45 m²", "Observação"]
ws.append(headers)
periods = [
    ("mai/2027", "Projeto executivo", .025), ("jun/2027", "Projeto executivo", .025),
    ("jul/2027", "Aprovações", .025), ("ago/2027", "Aprovações", .025),
    ("set/2027", "Base e estrutura", .05), ("out/2027", "Base e estrutura", .05), ("nov/2027", "Base e estrutura", .05),
    ("dez/2027", "Fechamentos", .05), ("jan/2028", "Fechamentos", .05),
    ("fev/2028", "Instalações", .05), ("mar/2028", "Instalações", .05),
    ("abr/2028", "Acabamentos", .05), ("mai/2028", "Acabamentos", .05),
    ("jun/2028", "Interiores", .075), ("jul/2028", "Interiores", .075),
    ("ago/2028", "Equipamentos e estoque", .15), ("set/2028", "Validações", .05), ("out/2028", "Abertura assistida", .10),
]
total_60, total_45 = 756001.80, 673366.80
for idx, (period, phase, share) in enumerate(periods, 1):
    note = "Evitar compra antecipada de equipamentos e produtos com validade" if idx < 16 else "Liberar somente após validações e cronograma real"
    ws.append([idx, period, phase, share, round(total_60 * share, 2), round(total_45 * share, 2), note])
ws.append(["TOTAL", "mai/2027–out/2028", "18 meses", "=SUM(D2:D19)", "=SUM(E2:E19)", "=SUM(F2:F19)", "Valores indicativos; substituir pelo fluxo contratado"])
gold, ink, ivory = "B8945E", "24312F", "F5F1EA"
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor=ink); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
for row in range(2, 21):
    ws.cell(row, 4).number_format = "0.0%"
    ws.cell(row, 5).number_format = 'R$ #,##0.00'
    ws.cell(row, 6).number_format = 'R$ #,##0.00'
    if row % 2 == 0:
        for col in range(1, 8): ws.cell(row, col).fill = PatternFill("solid", fgColor=ivory)
for col, width in {"A":9,"B":20,"C":27,"D":14,"E":18,"F":18,"G":55}.items(): ws.column_dimensions[col].width = width
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:G19"
ws[20][0].font = Font(bold=True, color=gold)
wb.save(path)
print(path)
